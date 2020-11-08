import openpyxl
import smtplib
import ssl

port = 465
sender_email = "oisin.farrell2000@gmail.com"
password = "xxxxxxxx"
context = ssl.create_default_context()
server = smtplib.SMTP_SSL("smtp.gmail.com", port, context=context)

## logging into your own email
server.login(sender_email, password)
receiver_email = "oisin.farrell2000@gmail.com"
breakfast_week = []
lunch_week = []
dinner_week = []
all_meals = []
all_possible_foods = []
groceries = []
wb = openpyxl.load_workbook('mealPlan.xlsx')
sheet = wb['Sheet1']
food_sheet = wb['Sheet2']
number_of_foods = food_sheet.max_column

for x in range(1, number_of_foods + 1):
    food = food_sheet.cell(row=1, column=x).value
    all_possible_foods.append(food)

for i in range(1, 8):
    breakfast = sheet.cell(row=1, column=i).value
    breakfast_week.append(breakfast)
    lunch = sheet.cell(row=2, column=i).value
    lunch_week.append(lunch)
    dinner = sheet.cell(row=3, column=i).value
    dinner_week.append(dinner)
all_meals.extend(breakfast_week)
all_meals.extend(lunch_week)
all_meals.extend(dinner_week)
for meal in all_meals:
    if meal in all_possible_foods:
        food_index = all_possible_foods.index(meal)
        index = 2
        ingredient = food_sheet.cell(row=index, column=food_index + 1).value
        while ingredient is not None:
            groceries.append(ingredient)
            index += 1
            ingredient = food_sheet.cell(row=index, column=food_index + 1).value

    else:
        print("{0} does not appear to be in the database ".format(meal))
        decision_to_add = input("Do you want to add it to the list [y/n]: ")
        if decision_to_add == "y":
            print(food_sheet.max_column)
            food_sheet.cell(row=1, column=15).value = 5
            food_sheet.cell(row=1, column=food_sheet.max_column + 1).value = str(meal)
            print(str(meal))
            print(food_sheet.cell(row=1, column=food_sheet.max_column + 1).value)

        #     new_ingredients = []
        #     ingredient_to_add = input("Enter an ingredient followed by return: ")
        #     while ingredient_to_add is not "":
        #         new_ingredients.append(ingredient_to_add)
        #
        # else:

grocery_list = []
for ing in groceries:
    if ing not in grocery_list:
        grocery_list.append(ing)

print("Grocery List: ")
emailed_list = "Grocery List:\n"
for z in grocery_list:
    emailed_list += z+"\n"
    print(z)

server.sendmail(sender_email, receiver_email, emailed_list)
print("email sent")
